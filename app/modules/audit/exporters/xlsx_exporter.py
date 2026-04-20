import io
from openpyxl import Workbook # type: ignore
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from .base import AuditLogExporter
from typing import Iterable, Any
# Importamos os mapeamentos que criamos anteriormente
from .base import ACTION_LABELS, COLUMN_HEADERS, STATUS_LABELS


class XLSXExporter(AuditLogExporter):
    def export(self, logs: Iterable[dict[str, Any]]) -> bytes:
        wb = Workbook()
        ws = wb.active
        ws.title = "Audit Logs"

        log_list = list(logs)

        display_columns = [
            "created_at",
            "username",
            "action",
            "target_username",
            "target_registration",
            "status",
            "ip_address",
            "message"
        ]

        if not log_list:
            buffer = io.BytesIO()
            wb.save(buffer)
            buffer.seek(0)
            return buffer.read()

        friendly_headers = [COLUMN_HEADERS.get(col, col) for col in display_columns]
        ws.append(friendly_headers)

        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )

        for col_num in range(1, len(friendly_headers) + 1):
            cell = ws.cell(row=1, column=col_num)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thin_border

        for log in log_list:
            row_data = [
                str(log.get("created_at")),
                log.get("username"),
                ACTION_LABELS.get(log.get("action"), log.get("action")),
                log.get("target_username") or "—",
                log.get("target_registration") or "—",
                STATUS_LABELS.get(log.get("status"), log.get("status")),
                log.get("ip_address"),
                log.get("message") or "—"
            ]

            sanitized_row = [
                (f"'{str(val)}" if str(val).startswith(("=", "+", "-", "@")) else val)
                for val in row_data
            ]
            ws.append(sanitized_row)

        ws.auto_filter.ref = ws.dimensions
        ws.freeze_panes = "A2"

        for col_num, column in enumerate(ws.columns, 1):
            max_length = 0
            column_letter = get_column_letter(col_num)

            for cell in column:
                try:
                    if cell.value:
                        length = len(str(cell.value))
                        if length > max_length:
                            max_length = length
                except:
                    pass

            adjusted_width = min(max(max_length + 2, 12), 50)
            ws.column_dimensions[column_letter].width = adjusted_width

        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer.read()