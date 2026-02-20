import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from .exporter_interface import AuditLogExporter
from typing import Iterable, Any


class XLSXExporter(AuditLogExporter):
    def export(self, logs: Iterable[dict[str, Any]]) -> bytes:
        wb = Workbook()
        ws = wb.active

        if ws is None:
            ws = wb.create_sheet("Audit Logs")
        else:
            ws.title = "Audit Logs"

        log_list = list(logs)

        if not log_list:
            buffer = io.BytesIO()
            wb.save(buffer)
            buffer.seek(0)
            return buffer.read()

        headers = list(log_list[0].keys())
        ws.append(headers)

        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(
            start_color="2C3E50",
            end_color="2C3E50",
            fill_type="solid"
        )
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )

        for col_num, _ in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thin_border

        for log in log_list:
            row_data = []
            for val in log.values():
                val_str = str(val) if val is not None else ""

                if val_str.startswith(("=", "+", "-", "@")):
                    val_str = "'" + val_str

                row_data.append(val_str)

            ws.append(row_data)

        if ws.auto_filter:
            ws.auto_filter.ref = ws.dimensions

        ws.freeze_panes = "A2"

        for col_num, column in enumerate(ws.columns, 1):
            max_length = 0
            column_letter = get_column_letter(col_num)

            if column[0].value:
                max_length = len(str(column[0].value))

            for cell in column[1:100]:
                try:
                    if cell.value and len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except Exception:
                    pass

            adjusted_width = (max_length + 2)
            final_width = min(max(adjusted_width, 10), 60)
            ws.column_dimensions[column_letter].width = final_width

        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer.read()
